from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Query, Depends
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, update, and_, or_, desc, delete
from sqlalchemy.orm import undefer
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, field_validator
from typing import List, Literal, Optional, Union, Any
import uuid
from datetime import datetime, timezone, timedelta
import shutil
import jwt
import bcrypt
import httpx
import secrets
import random
import io

import models_sqlalchemy as sql_models
from auth_utils import get_user_product_ids, get_user_depot_ids, build_product_filter, ensure_transporter_access, build_transporter_filter, normalize_mobile, get_current_user, check_permission, get_download_user, create_download_token, DOWNLOAD_TOKEN_TTL_SECONDS
from database import engine, Base, get_db, init_db, AsyncSessionLocal
from config import PERMISSION_DEFAULTS
from routes.db_compat import db
from tenant import tenant_filter, tenant_id_for_current_user, PLATFORM_TENANT_ID

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

DATABASE_URL = os.environ.get('MYSQL_URL')

UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

JWT_SECRET = os.environ.get('JWT_SECRET')
if not JWT_SECRET:
    raise ValueError("JWT_SECRET environment variable is required")
JWT_ALGORITHM = "HS256"

MSG91_AUTHKEY = os.environ.get('MSG91_AUTHKEY', '')
MSG91_TEMPLATE_ID = os.environ.get('MSG91_TEMPLATE_ID', '')
MSG91_DLT_TE_ID = os.environ.get('MSG91_DLT_TE_ID', '')
MSG91_SENDER_ID = "INFOET"
OTP_EXPIRY_SECONDS = 120
MAX_OTP_ATTEMPTS = 5
# Echoing the OTP back in the HTTP response defeats SMS delivery as an auth
# factor, so it is opt-in for local development only. Never enable in production.
EXPOSE_OTP_IN_RESPONSE = os.environ.get('EXPOSE_OTP_IN_RESPONSE', '').strip().lower() in {'1', 'true', 'yes'}

COUNTRY_CODES = {
    "IN": {"code": "91", "name": "India", "flag": "🇮🇳"},
    "NP": {"code": "977", "name": "Nepal", "flag": "🇳🇵"},
    "BD": {"code": "880", "name": "Bangladesh", "flag": "🇧🇩"},
    "VN": {"code": "84", "name": "Vietnam", "flag": "🇻🇳"},
    "BT": {"code": "975", "name": "Bhutan", "flag": "🇧🇹"},
    "AE": {"code": "971", "name": "UAE", "flag": "🇦🇪"},
}

app = FastAPI()
api_router = APIRouter(prefix="/api/v1")
security = HTTPBearer()


def generate_otp(length: int = 6) -> str:
    return ''.join([str(random.randint(0, 9)) for _ in range(length)])


def otp_response(payload: dict, otp_code: str) -> dict:
    if EXPOSE_OTP_IN_RESPONSE:
        return {**payload, "demo_otp": otp_code}
    return payload


def ensure_otp_attempts_remaining(otp_record):
    if otp_record.attempts >= MAX_OTP_ATTEMPTS:
        raise HTTPException(status_code=429, detail="Too many failed attempts. Please request a new OTP.")


async def reject_invalid_otp(otp_record):
    """Record a failed OTP guess and raise 401. Never returns."""
    async with AsyncSessionLocal() as session:
        await session.execute(
            update(sql_models.OTP).where(sql_models.OTP.id == otp_record.id).values(attempts=otp_record.attempts + 1)
        )
        await session.commit()
    remaining = MAX_OTP_ATTEMPTS - otp_record.attempts - 1
    raise HTTPException(status_code=401, detail=f"Invalid OTP. {remaining} attempts remaining.")


def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())



def create_token(user_data: dict) -> str:
    payload = {
        "user_id": user_data["id"],
        "mobile": user_data["mobile"],
        "role": user_data["role"],
        "name": user_data["name"],
        "tenant_id": user_data.get("tenant_id"),
        "exp": datetime.now(timezone.utc).timestamp() + 86400 * 7
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)



async def send_otp_via_msg91(mobile: str, country_code: str, otp_code: str) -> dict:
    full_mobile = f"{country_code}{mobile}"
    expiry_mins = OTP_EXPIRY_SECONDS // 60 or 1
    url = "https://api.msg91.com/api/v2/sendsms"
    headers = {"authkey": MSG91_AUTHKEY, "Content-Type": "application/json"}
    message = f"{otp_code} is the OTP for accessing your infoEIGHT account. Please do not share it with anyone. OTP will be valid for {expiry_mins} mins."
    payload = {
        "sender": MSG91_SENDER_ID,
        "route": "4",
        "country": country_code,
        "DLT_TE_ID": MSG91_DLT_TE_ID,
        "sms": [{"message": message, "to": [mobile]}],
    }
    logging.info(f"Sending OTP to {full_mobile} via MSG91 SMS API")
    async with httpx.AsyncClient() as client:
        try:
            response = await client.post(url, json=payload, headers=headers, timeout=10)
            return response.json()
        except Exception:
            return {"type": "success", "message": "OTP sent (demo mode)"}


# "registration" is deliberately absent -- there is no self-registration flow,
# so an OTP issued for that purpose would have nothing to authorise. Anything
# outside this set is rejected with a 422 before an SMS is sent.
OtpPurpose = Literal["login", "reset_password", "first_time_setup"]


class SendOTPRequest(BaseModel):
    mobile: str
    country_code: str = "91"
    purpose: OtpPurpose = "login"


class VerifyOTPRequest(BaseModel):
    mobile: str
    country_code: str = "91"
    otp_code: str
    purpose: OtpPurpose = "login"
    tenant: Optional[str] = None


class LoginWithPasswordRequest(BaseModel):
    mobile: str
    country_code: str = "91"
    password: str


class LoginWithOTPRequest(BaseModel):
    mobile: str
    country_code: str = "91"
    tenant: Optional[str] = None


class FirstTimeSetupRequest(BaseModel):
    mobile: str
    country_code: str = "91"
    otp_code: str
    new_password: str
    tenant: Optional[str] = None


class AdminCreateUserRequest(BaseModel):
    name: str
    mobile: str
    country_code: str = "91"
    role: str
    company_id: Optional[str] = None
    email: Optional[str] = None
    depot_id: Optional[str] = None
    employee_id: Optional[str] = None
    assigned_products: List[str] = []
    assigned_depots: List[str] = []


class UpdateUserProductsRequest(BaseModel):
    assigned_products: List[str] = []


class ResetPasswordRequest(BaseModel):
    mobile: str
    country_code: str = "91"
    otp_code: str
    new_password: str
    tenant: Optional[str] = None


class TokenResponse(BaseModel):
    token: str
    user: dict


class UserLogin(BaseModel):
    mobile: str
    country_code: str = "91"
    password: str
    tenant: Optional[str] = None


async def _tenant_id_by_slug(tenant: Optional[str]) -> Optional[str]:
    """Resolve an optional tenant slug to its id (None when not provided)."""
    if not tenant:
        return None
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(sql_models.Tenant.id).where(sql_models.Tenant.slug == tenant)
        )
        tenant_id = result.scalar_one_or_none()
    if not tenant_id:
        raise HTTPException(status_code=404, detail="Unknown tenant")
    return tenant_id


async def _find_user_by_mobile(
    mobile: str,
    country_code: str,
    tenant_id: Optional[str] = None,
):
    """Resolve a user by mobile number, disambiguating across tenants.

    uk_mobile is per-tenant, so the same number can exist in several tenants.
    When it does and no tenant is given, the caller must supply one -- the
    login page then asks for the tenant slug. Mirrors the legacy lookup, which
    also tried the digits-only and prefixed spellings of the number.
    """
    full_mobile = normalize_mobile(mobile, country_code)
    spellings = {full_mobile}
    legacy = ''.join(filter(str.isdigit, str(mobile)))
    if len(legacy) == 10:
        spellings.add(legacy)
        if f"{country_code}{legacy}" != full_mobile:
            spellings.add(f"{country_code}{legacy}")
    async with AsyncSessionLocal() as session:
        stmt = select(sql_models.User).where(sql_models.User.mobile.in_(spellings))
        if tenant_id:
            stmt = stmt.where(sql_models.User.tenant_id == tenant_id)
        result = await session.execute(stmt.limit(20))
        users = result.scalars().all()
    if not users:
        return None
    if len(users) == 1:
        return users[0]
    if tenant_id:
        return next(
            (u for u in users if (u.get("tenant_id") if hasattr(u, "get") else u.tenant_id) == tenant_id),
            None,
        )
    raise HTTPException(
        status_code=401,
        detail="This mobile belongs to more than one workspace. Please provide your tenant slug.",
    )


@api_router.get("/country-codes")
async def get_country_codes():
    return list(COUNTRY_CODES.values())


@api_router.post("/otp/send")
async def send_otp(request: SendOTPRequest):
    full_mobile = f"{request.country_code}{request.mobile}"
    five_minutes_ago = datetime.now(timezone.utc) - timedelta(minutes=5)
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(func.count(sql_models.OTP.id)).where(
                and_(sql_models.OTP.mobile == full_mobile, sql_models.OTP.created_at >= five_minutes_ago)
            )
        )
        recent_otps = result.scalar_one()
    if recent_otps >= 3:
        raise HTTPException(status_code=429, detail="Too many OTP requests. Please wait 5 minutes.")
    otp_code = generate_otp()
    expires_at = datetime.now(timezone.utc) + timedelta(seconds=OTP_EXPIRY_SECONDS)
    otp_doc = sql_models.OTP(
        id=str(uuid.uuid4()),
        mobile=full_mobile,
        country_code=request.country_code,
        otp_code=otp_code,
        purpose=request.purpose,
        verified=False,
        attempts=0,
        created_at=datetime.now(timezone.utc),
        expires_at=expires_at,
    )
    async with AsyncSessionLocal() as session:
        session.add(otp_doc)
        await session.commit()
    await send_otp_via_msg91(request.mobile, request.country_code, otp_code)
    return otp_response({
        "success": True,
        "message": f"OTP sent to +{full_mobile}",
        "expires_in_seconds": OTP_EXPIRY_SECONDS,
    }, otp_code)


@api_router.post("/otp/verify")
async def verify_otp(request: VerifyOTPRequest):
    full_mobile = f"{request.country_code}{request.mobile}"
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(sql_models.OTP).where(
                and_(
                    sql_models.OTP.mobile == full_mobile,
                    sql_models.OTP.purpose == request.purpose,
                    sql_models.OTP.verified == False,
                )
            ).order_by(desc(sql_models.OTP.created_at)).limit(1)
        )
        otp_record = result.scalar_one_or_none()
    if not otp_record:
        raise HTTPException(status_code=404, detail="No pending OTP found. Please request a new one.")
    if datetime.utcnow() > otp_record.expires_at:
        raise HTTPException(status_code=410, detail="OTP has expired. Please request a new one.")
    ensure_otp_attempts_remaining(otp_record)
    if otp_record.otp_code != request.otp_code:
        await reject_invalid_otp(otp_record)
    async with AsyncSessionLocal() as session:
        await session.execute(
            update(sql_models.OTP).where(sql_models.OTP.id == otp_record.id).values(verified=True)
        )
        await session.commit()
    verification_token = secrets.token_urlsafe(32)
    return {"success": True, "message": "OTP verified successfully", "verification_token": verification_token}


@api_router.post("/otp/resend")
async def resend_otp(request: SendOTPRequest):
    return await send_otp(request)


# Self-registration is deliberately not offered. Accounts exist only where an
# Admin or Management user created them via POST /api/admin/users; the account
# holder then sets their own password through /auth/first-time-setup.
# The removed /auth/register took `role` straight from the request body, so any
# caller who could receive an SMS could mint themselves a Management account.


@api_router.post("/auth/login")
async def login(data: UserLogin):
    tenant_id = await _tenant_id_by_slug(data.tenant)
    user = await _find_user_by_mobile(data.mobile, data.country_code, tenant_id)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid mobile number or password")
    if not user.password_set:
        otp_code = generate_otp()
        expires_at = datetime.now(timezone.utc) + timedelta(seconds=OTP_EXPIRY_SECONDS)
        otp_doc = sql_models.OTP(
            id=str(uuid.uuid4()),
            mobile=full_mobile,
            country_code=data.country_code,
            otp_code=otp_code,
            purpose="first_time_setup",
            verified=False,
            attempts=0,
            created_at=datetime.now(timezone.utc),
            expires_at=expires_at,
        )
        async with AsyncSessionLocal() as session:
            session.add(otp_doc)
            await session.commit()
        await send_otp_via_msg91(data.mobile, data.country_code, otp_code)
        return otp_response({
            "first_time_login": True,
            "message": f"OTP sent to +{full_mobile}. Please set your password.",
            "expires_in_seconds": OTP_EXPIRY_SECONDS,
        }, otp_code)
    if not verify_password(data.password, user.password):
        raise HTTPException(status_code=401, detail="Invalid mobile number or password")
    user_response = {
        "id": user.id,
        "tenant_id": user.tenant_id,
        "name": user.name,
        "mobile": user.mobile,
        "country_code": user.country_code,
        "role": user.role,
        "email": user.email,
        "depot_id": user.depot_id,
        "company_id": user.company_id,
        "otp_verified": user.otp_verified,
        "password_set": user.password_set,
        "is_master_admin": user.is_master_admin,
        "employee_id": user.employee_id,
    }
    token = create_token(user.__dict__)
    return {"token": token, "user": user_response}


@api_router.post("/auth/first-time-setup", response_model=TokenResponse)
async def first_time_setup(data: FirstTimeSetupRequest):
    tenant_id = await _tenant_id_by_slug(data.tenant)
    user = await _find_user_by_mobile(data.mobile, data.country_code, tenant_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.password_set:
        raise HTTPException(status_code=400, detail="Password already set. Please use regular login.")
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(sql_models.OTP).where(
                and_(
                    sql_models.OTP.mobile == full_mobile,
                    sql_models.OTP.purpose == "first_time_setup",
                    sql_models.OTP.verified == False,
                )
            ).order_by(desc(sql_models.OTP.created_at)).limit(1)
        )
        otp_record = result.scalar_one_or_none()
    if not otp_record:
        raise HTTPException(status_code=404, detail="No pending OTP found. Please try logging in again.")
    if datetime.utcnow() > otp_record.expires_at:
        raise HTTPException(status_code=410, detail="OTP has expired. Please try logging in again.")
    ensure_otp_attempts_remaining(otp_record)
    if otp_record.otp_code != data.otp_code:
        await reject_invalid_otp(otp_record)
    async with AsyncSessionLocal() as session:
        await session.execute(
            update(sql_models.User)
            .where(sql_models.User.id == user.id)
            .values(
                password=hash_password(data.new_password),
                password_set=True,
                otp_verified=True,
            )
        )
        await session.commit()
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(sql_models.User).where(sql_models.User.id == user.id))
        updated_user = result.scalar_one_or_none()
    user_response = {
        "id": updated_user.id,
        "tenant_id": updated_user.tenant_id,
        "name": updated_user.name,
        "mobile": updated_user.mobile,
        "country_code": updated_user.country_code,
        "role": updated_user.role,
        "email": updated_user.email,
        "depot_id": updated_user.depot_id,
        "otp_verified": True,
        "password_set": True,
        "is_master_admin": updated_user.is_master_admin,
        "employee_id": updated_user.employee_id,
    }
    token = create_token(updated_user.__dict__)
    return {"token": token, "user": user_response}


@api_router.post("/auth/login-otp")
async def login_with_otp(data: LoginWithOTPRequest):
    tenant_id = await _tenant_id_by_slug(data.tenant)
    user = await _find_user_by_mobile(data.mobile, data.country_code, tenant_id)
    if not user:
        raise HTTPException(status_code=404, detail="No account found with this mobile number")
    otp_request = SendOTPRequest(mobile=data.mobile, country_code=data.country_code, purpose="login")
    return await send_otp(otp_request)


@api_router.post("/auth/login-otp/verify", response_model=TokenResponse)
async def verify_login_otp(request: VerifyOTPRequest):
    full_mobile = f"{request.country_code}{request.mobile}"
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(sql_models.OTP).where(
                and_(
                    sql_models.OTP.mobile == full_mobile,
                    sql_models.OTP.purpose == "login",
                    sql_models.OTP.verified == False,
                )
            ).order_by(desc(sql_models.OTP.created_at)).limit(1)
        )
        otp_record = result.scalar_one_or_none()
    if not otp_record:
        raise HTTPException(status_code=404, detail="No pending OTP found")
    if datetime.utcnow() > otp_record.expires_at:
        raise HTTPException(status_code=410, detail="OTP has expired")
    ensure_otp_attempts_remaining(otp_record)
    if otp_record.otp_code != request.otp_code:
        await reject_invalid_otp(otp_record)
    async with AsyncSessionLocal() as session:
        await session.execute(
            update(sql_models.OTP).where(sql_models.OTP.id == otp_record.id).values(verified=True)
        )
        await session.commit()
    tenant_id = await _tenant_id_by_slug(request.tenant)
    user = await _find_user_by_mobile(request.mobile, request.country_code, tenant_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user_response = {
        "id": user.id,
        "tenant_id": user.tenant_id,
        "name": user.name,
        "mobile": user.mobile,
        "country_code": user.country_code,
        "role": user.role,
        "email": user.email,
        "depot_id": user.depot_id,
        "company_id": user.company_id,
        "otp_verified": user.otp_verified,
        "password_set": user.password_set,
        "is_master_admin": user.is_master_admin,
        "employee_id": user.employee_id,
    }
    token = create_token(user.__dict__)
    async with AsyncSessionLocal() as session:
        await session.execute(
            update(sql_models.OTP).where(sql_models.OTP.mobile == full_mobile, sql_models.OTP.purpose == "login").values(verified=True)
        )
        await session.commit()
    return {"token": token, "user": user_response}


@api_router.post("/auth/forgot-password")
async def forgot_password(data: LoginWithOTPRequest):
    tenant_id = await _tenant_id_by_slug(data.tenant)
    user = await _find_user_by_mobile(data.mobile, data.country_code, tenant_id)
    if not user:
        raise HTTPException(status_code=404, detail="No account found with this mobile number")
    otp_request = SendOTPRequest(mobile=data.mobile, country_code=data.country_code, purpose="reset_password")
    return await send_otp(otp_request)


@api_router.post("/auth/reset-password")
async def reset_password(data: ResetPasswordRequest):
    full_mobile = normalize_mobile(data.mobile, data.country_code)
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(sql_models.OTP).where(
                and_(
                    sql_models.OTP.mobile == full_mobile,
                    sql_models.OTP.purpose == "reset_password",
                    sql_models.OTP.verified == False,
                )
            ).order_by(desc(sql_models.OTP.created_at)).limit(1)
        )
        otp_record = result.scalar_one_or_none()
    if not otp_record:
        raise HTTPException(status_code=404, detail="No pending OTP found. Request password reset first.")
    if datetime.utcnow() > otp_record.expires_at:
        raise HTTPException(status_code=410, detail="OTP has expired")
    ensure_otp_attempts_remaining(otp_record)
    if otp_record.otp_code != data.otp_code:
        await reject_invalid_otp(otp_record)
    tenant_id = await _tenant_id_by_slug(data.tenant)
    user = await _find_user_by_mobile(data.mobile, data.country_code, tenant_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    async with AsyncSessionLocal() as session:
        await session.execute(
            update(sql_models.User).where(sql_models.User.id == user.id).values(password=hash_password(data.new_password))
        )
        await session.commit()
    async with AsyncSessionLocal() as session:
        await session.execute(
            update(sql_models.OTP).where(sql_models.OTP.mobile == full_mobile, sql_models.OTP.purpose == "reset_password").values(verified=True)
        )
        await session.commit()
    return {"success": True, "message": "Password reset successfully. Please login with your new password."}


@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return current_user


@api_router.post("/auth/download-token")
async def issue_download_token(current_user: dict = Depends(get_current_user)):
    """Mint a short-lived, download-scoped token.

    The browser loads file and export URLs itself (`<img src>`, `window.open`)
    and cannot attach an Authorization header, so those URLs carry this token as
    `?t=`. It is useless anywhere else and expires in 30 minutes, which keeps a
    full-privilege token out of access logs and browser history.
    """
    return {
        "token": create_download_token(current_user["id"]),
        "expires_in_seconds": DOWNLOAD_TOKEN_TTL_SECONDS,
    }


@api_router.post("/admin/users")
async def admin_create_user(data: AdminCreateUserRequest, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in {"Management", "Admin"}:
        raise HTTPException(status_code=403, detail="Only Admin or Management can create users")
    full_mobile = normalize_mobile(data.mobile, data.country_code)
    mobile_digits = ''.join(filter(str.isdigit, str(data.mobile or "")))
    query_conditions = [sql_models.User.mobile == full_mobile]
    if mobile_digits:
        if mobile_digits != full_mobile:
            query_conditions.append(sql_models.User.mobile == mobile_digits)
        if len(mobile_digits) == 10 and f"{data.country_code}{mobile_digits}" != full_mobile:
            query_conditions.append(sql_models.User.mobile == f"{data.country_code}{mobile_digits}")
    async with AsyncSessionLocal() as session:
        # The OR spans up to three spellings of the same number. Each is unique on
        # its own, but they can match different rows, so take the first match
        # rather than assert there is exactly one. Scoped to the caller's tenant:
        # the same mobile may exist in another tenant.
        stmt = select(sql_models.User).where(or_(*query_conditions)).limit(1)
        tfilter = tenant_filter(sql_models.User)
        if tfilter is not None:
            stmt = stmt.where(tfilter)
        result = await session.execute(stmt)
        existing = result.scalars().first()
    if existing:
        existing_name = existing.name or "This user"
        raise HTTPException(status_code=400, detail=f"{existing_name} has already been assigned this number")
    user_id = str(uuid.uuid4())
    user_doc = sql_models.User(
        id=user_id,
        tenant_id=tenant_id_for_current_user(current_user),
        employee_id=data.employee_id,
        name=data.name,
        mobile=full_mobile,
        country_code=data.country_code,
        password="",
        password_set=False,
        role=data.role,
        email=data.email,
        depot_id=data.depot_id,
        company_id=None if data.role == "Transporter" else current_user.get("company_id"),
        assigned_products=data.assigned_products,
        assigned_depots=data.assigned_depots,
        excluded_products=[],
        excluded_depots=[],
        otp_verified=False,
        created_by=current_user.get("id"),
        created_at=datetime.now(timezone.utc),
    )
    if data.role == "Transporter":
        transporter_id = data.transporter_id or data.company_id
        if transporter_id:
            async with AsyncSessionLocal() as session:
                stmt = select(sql_models.Transporter).where(sql_models.Transporter.id == transporter_id)
                tfilter = tenant_filter(sql_models.Transporter)
                if tfilter is not None:
                    stmt = stmt.where(tfilter)
                result = await session.execute(stmt)
                transporter = result.scalar_one_or_none()
            if transporter:
                user_doc.transporter_id = transporter_id
                user_doc.transporter_name = transporter.name or ""
                transporter.users = transporter.users or []
                transporter.users.append({"id": user_id, "created_at": user_doc.created_at.isoformat() if user_doc.created_at else None})
                async with AsyncSessionLocal() as session:
                    session.add(transporter)
                    await session.commit()
    if data.role != "Transporter" and not current_user.get("company_id") and not current_user.get("is_master_admin"):
        raise HTTPException(400, "User must belong to a company")
    async with AsyncSessionLocal() as session:
        session.add(user_doc)
        await session.commit()

    # Sync the reverse linkage on the employee record (Phase 3).
    if data.employee_id:
        async with AsyncSessionLocal() as session:
            emp = (await session.execute(
                select(sql_models.Employee).where(sql_models.Employee.id == data.employee_id)
            )).scalar_one_or_none()
            if emp:
                emp.user_id = user_id
                emp.login_enabled = True
                await session.commit()
    user_response = {
        "id": user_doc.id,
        "name": user_doc.name,
        "mobile": user_doc.mobile,
        "country_code": user_doc.country_code,
        "role": user_doc.role,
        "email": user_doc.email,
        "depot_id": user_doc.depot_id,
        "company_id": user_doc.company_id,
        "transporter_id": user_doc.transporter_id,
        "transporter_name": user_doc.transporter_name,
        "assigned_products": user_doc.assigned_products,
        "assigned_depots": user_doc.assigned_depots,
        "otp_verified": user_doc.otp_verified,
        "password_set": user_doc.password_set,
    }
    return {"success": True, "message": f"User created. They can login with mobile +{full_mobile}", "user": user_response}


@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    original_filename = file.filename
    import re
    safe_filename = re.sub(r'[^\w\-_\.]', '_', original_filename)
    unique_prefix = str(uuid.uuid4())[:8]
    file_name = f"{unique_prefix}_{safe_filename}"
    tenant_dir = UPLOAD_DIR / (current_user.get("tenant_id") or PLATFORM_TENANT_ID)
    tenant_dir.mkdir(exist_ok=True)
    file_path = tenant_dir / file_name
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return {"file_id": file_name, "filename": original_filename, "original_name": original_filename}


@api_router.get("/uploads/{file_id}")
async def get_file(file_id: str, current_user: dict = Depends(get_download_user)):
    # Tenant-isolated storage with a legacy-root fallback so files uploaded
    # before Phase 0 keep working without any file moves.
    candidate = None
    if current_user:
        candidate = UPLOAD_DIR / (current_user.get("tenant_id") or PLATFORM_TENANT_ID) / file_id
        if not candidate.exists():
            candidate = None
    if candidate is None:
        candidate = UPLOAD_DIR / file_id
    file_path = candidate
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    original_name = file_id.split('_', 1)[1] if '_' in file_id else file_id
    ext = file_id.rsplit('.', 1)[-1].lower() if '.' in file_id else ''
    content_types = {
        'pdf': 'application/pdf',
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'png': 'image/png',
        'gif': 'image/gif',
    }
    media_type = content_types.get(ext, 'application/octet-stream')
    return FileResponse(file_path, media_type=media_type, filename=original_name)


# Depot inventory movements live in routes/liftings.py (update_depot_inventory).
# An unused copy of that logic sat here and drifted out of sync with it.


@api_router.get("/users")
async def get_users(current_user: dict = Depends(get_current_user)):
    async with AsyncSessionLocal() as session:
        stmt = select(sql_models.User).options(undefer(sql_models.User.password))
        tfilter = tenant_filter(sql_models.User)
        if tfilter is not None:
            stmt = stmt.where(tfilter)
        result = await session.execute(stmt)
        users = result.scalars().all()
        user_list = []
        for u in users:
            ud = {
                "id": u.id,
                "name": u.name,
                "mobile": u.mobile,
                "country_code": u.country_code,
                "role": u.role,
                "email": u.email,
                "depot_id": u.depot_id,
                "company_id": u.company_id,
                "transporter_id": u.transporter_id,
                "assigned_products": u.assigned_products,
                "assigned_depots": u.assigned_depots,
                "excluded_products": u.excluded_products,
                "excluded_depots": u.excluded_depots,
                "otp_verified": u.otp_verified,
                "password_set": u.password_set,
                "is_master_admin": u.is_master_admin,
                "employee_id": u.employee_id,
                "created_by": u.created_by,
                "created_at": u.created_at.isoformat() if u.created_at else None,
            }
            user_list.append(ud)
        return user_list


class UpdateUserRequest(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    email: Optional[str] = None
    depot_id: Optional[str] = None


@api_router.put("/users/{user_id}")
async def update_user(user_id: str, data: UpdateUserRequest, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "Management":
        raise HTTPException(status_code=403, detail="Only Management can update users")
    async with AsyncSessionLocal() as session:
        stmt = select(sql_models.User).where(sql_models.User.id == user_id)
        tfilter = tenant_filter(sql_models.User)
        if tfilter is not None:
            stmt = stmt.where(tfilter)
        result = await session.execute(stmt)
        user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.is_master_admin:
        raise HTTPException(status_code=403, detail="Cannot modify Master Admin")
    update_data = {}
    if data.name is not None:
        update_data["name"] = data.name
    if data.role is not None:
        update_data["role"] = data.role
    if data.email is not None:
        update_data["email"] = data.email
    if data.depot_id is not None:
        update_data["depot_id"] = data.depot_id
    if update_data:
        async with AsyncSessionLocal() as session:
            await session.execute(
                update(sql_models.User).where(sql_models.User.id == user_id).values(**update_data)
            )
            await session.commit()
    return {"success": True, "message": "User updated successfully"}


@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "Management":
        raise HTTPException(status_code=403, detail="Only Management can delete users")
    async with AsyncSessionLocal() as session:
        stmt = select(sql_models.User).where(sql_models.User.id == user_id)
        tfilter = tenant_filter(sql_models.User)
        if tfilter is not None:
            stmt = stmt.where(tfilter)
        result = await session.execute(stmt)
        user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.is_master_admin:
        raise HTTPException(status_code=403, detail="Cannot delete Master Admin")
    if user_id == current_user.get("id"):
        raise HTTPException(status_code=403, detail="Cannot delete your own account")
    async with AsyncSessionLocal() as session:
        await session.delete(user)
        await session.commit()
    return {"message": "User deleted"}


# ============ EXPORT ROUTES ============

@api_router.get("/export/liftings")
async def export_liftings(
    format: str = "excel",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    product_id: Optional[str] = None,
    delivery_order_id: Optional[str] = None,
    vehicle_id: Optional[str] = None,
    transporter_name: Optional[str] = None,
    loading_point_id: Optional[str] = None,
    unloading_point_id: Optional[str] = None,
    unloading_status: Optional[str] = None,
    lifting_type: Optional[str] = None,
    current_user: dict = Depends(get_download_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    conditions = []
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to
        conditions.append({"date_of_loading": date_query})
    if product_id:
        conditions.append({"product_id": product_id})
    if delivery_order_id:
        conditions.append({"delivery_order_id": delivery_order_id})
    if vehicle_id:
        conditions.append({"vehicle_id": vehicle_id})
    if transporter_name:
        conditions.append({"transporter_name": {"$regex": transporter_name, "$options": "i"}})
    if loading_point_id:
        conditions.append({"loading_point_id": loading_point_id})
    if unloading_point_id:
        conditions.append({"unloading_point_id": unloading_point_id})
    if unloading_status:
        conditions.append({"unloading_status": unloading_status})
    if lifting_type:
        conditions.append({"lifting_type": lifting_type})

    liftings = await db.liftings.find({"$and": conditions} if conditions else {}, {"_id": 0}).to_list(10000)

    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Liftings Report"
        headers = ["Lifting No", "Date", "Type", "Product", "Quantity (MT)", "Vehicle",
                   "Transporter", "Driver", "From", "To", "Status", "Net Weight (MT)"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for row, lifting in enumerate(liftings, 2):
            ws.cell(row=row, column=1, value=lifting.get("lifting_no", ""))
            ws.cell(row=row, column=2, value=lifting.get("date_of_loading", ""))
            ws.cell(row=row, column=3, value=lifting.get("lifting_type", ""))
            ws.cell(row=row, column=4, value=lifting.get("product_name", ""))
            ws.cell(row=row, column=5, value=lifting.get("quantity_mt", 0))
            ws.cell(row=row, column=6, value=lifting.get("vehicle_number", ""))
            ws.cell(row=row, column=7, value=lifting.get("transporter_name", ""))
            ws.cell(row=row, column=8, value=lifting.get("driver_name", ""))
            ws.cell(row=row, column=9, value=lifting.get("loading_point_name", ""))
            ws.cell(row=row, column=10, value=lifting.get("unloading_point_name", ""))
            ws.cell(row=row, column=11, value=lifting.get("unloading_status", ""))
            ws.cell(row=row, column=12, value=lifting.get("net_weight_mt", 0))
            for col in range(1, 13):
                ws.cell(row=row, column=col).border = thin_border
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"liftings_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"})
    return {"error": "Unsupported format"}


@api_router.get("/export/inventory")
async def export_inventory(format: str = "excel", current_user: dict = Depends(get_download_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    inventory = await db.depot_inventory.find({}, {"_id": 0}).to_list(10000)
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Report"
        headers = ["Depot Name", "Product Name", "Quantity (MT)", "Last Updated"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for row, item in enumerate(inventory, 2):
            ws.cell(row=row, column=1, value=item.get("depot_name", ""))
            ws.cell(row=row, column=2, value=item.get("product_name", ""))
            ws.cell(row=row, column=3, value=item.get("quantity_mt", 0))
            ws.cell(row=row, column=4, value=item.get("updated_at", ""))
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = thin_border
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"})
    return {"error": "Unsupported format"}


@api_router.get("/export/delivery-orders")
async def export_delivery_orders(format: str = "excel", status: Optional[str] = None, current_user: dict = Depends(get_download_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    query = {}
    if status:
        query["status"] = status
    orders = await db.delivery_orders.find(query, {"_id": 0}).to_list(10000)
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Delivery Orders"
        headers = ["DO Number", "Date", "Product", "Company", "Total Qty (MT)",
                   "Remaining Qty (MT)", "To Depot", "Status"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for row, order in enumerate(orders, 2):
            ws.cell(row=row, column=1, value=order.get("do_order_no", ""))
            ws.cell(row=row, column=2, value=order.get("do_date", ""))
            ws.cell(row=row, column=3, value=order.get("product_name", ""))
            ws.cell(row=row, column=4, value=order.get("from_company_name", ""))
            ws.cell(row=row, column=5, value=order.get("total_quantity_mt", 0))
            ws.cell(row=row, column=6, value=order.get("remaining_quantity_mt", 0))
            ws.cell(row=row, column=7, value=order.get("to_depot_name", ""))
            ws.cell(row=row, column=8, value=order.get("status", ""))
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"delivery_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"})
    return {"error": "Unsupported format"}


@api_router.get("/export/users")
async def export_users(format: str = "excel", current_user: dict = Depends(get_download_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(10000)
    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Users Report"
        headers = ["S.No", "Name", "Mobile", "Country Code", "Email", "Role",
                   "Depot ID", "OTP Verified", "Password Set", "Created At"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for row_idx, user in enumerate(users, 2):
            row_data = [row_idx - 1, user.get('name', ''), user.get('mobile', ''),
                        user.get('country_code', '91'), user.get('email', ''),
                        user.get('role', ''), user.get('depot_id', ''),
                        'Yes' if user.get('otp_verified') else 'No',
                        'Yes' if user.get('password_set') else 'No',
                        user.get('created_at', '')[:19] if user.get('created_at') else '']
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"})
    return {"error": "Unsupported format"}


@api_router.get("/purchase-orders/{order_id}/statement/export")
async def export_purchase_order_statement(order_id: str, format: str = "excel", current_user: dict = Depends(get_download_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    order = await db.purchase_orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Purchase Order not found")

    liftings = await db.liftings.find(
        {"purchase_order_id": order_id, "unloading_status": {"$ne": "Rejected"}},
        {"_id": 0}
    ).sort("date_of_loading", -1).to_list(1000)

    pickups = await db.pickups.find(
        {"purchase_order_id": order_id, "status": {"$in": ["verified", "weightment_done", "final_verified"]}},
        {"_id": 0}
    ).sort("verified_at", -1).to_list(1000)

    transactions = []
    for p in pickups:
        transactions.append({
            "date": p.get("verified_at") or p.get("date"),
            "type": "Pickup",
            "reference_no": p.get("purchase_order_no"),
            "vehicle": p.get("truck_number"),
            "quantity": p.get("loaded_weight_mt") or p.get("weight_mt") or 0,
            "status": p.get("status", "")
        })

    lifted_quantity = sum(l.get("net_weight_mt") or l.get("quantity_mt", 0) for l in liftings)
    transactions.sort(key=lambda x: x.get("date") or "", reverse=True)

    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "PO Statement"
        headers = ["Date", "Type", "Reference", "Vehicle", "Quantity", "Status"]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for row, tx in enumerate(transactions, 2):
            ws.cell(row=row, column=1, value=tx.get("date", ""))
            ws.cell(row=row, column=2, value=tx.get("type", ""))
            ws.cell(row=row, column=3, value=tx.get("reference_no", ""))
            ws.cell(row=row, column=4, value=tx.get("vehicle", ""))
            ws.cell(row=row, column=5, value=tx.get("quantity", 0))
            ws.cell(row=row, column=6, value=tx.get("status", ""))
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"PO_Statement_{order.get('po_number', 'N/A')}.xlsx"
        return StreamingResponse(buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"})
    if format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            rightMargin=15 * mm, leftMargin=15 * mm,
            topMargin=15 * mm, bottomMargin=15 * mm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle', parent=styles['Title'], fontSize=16, spaceAfter=6,
        )
        subtitle_style = ParagraphStyle(
            'SubtitleStyle', parent=styles['Normal'], fontSize=10,
            textColor=colors.grey, spaceAfter=12,
        )
        header_style = ParagraphStyle(
            'HeaderStyle', parent=styles['Normal'], fontSize=9,
            textColor=colors.white, alignment=1,
        )
        body_style = ParagraphStyle(
            'BodyStyle', parent=styles['Normal'], fontSize=8.5,
        )

        po_no = order.get('po_number', 'N/A')
        company_name = order.get('to_company_name') or order.get('from_company_name') or ''
        story = [
            Paragraph(f"PO Statement - {po_no}", title_style),
            Paragraph(f"{company_name}<br/>Total Lifted Quantity: {lifted_quantity}", subtitle_style),
        ]

        headers = ["Date", "Type", "Reference", "Vehicle", "Quantity", "Status"]
        data = [[Paragraph(h, header_style) for h in headers]]
        for tx in transactions:
            data.append([
                Paragraph(str(tx.get("date", "") or ""), body_style),
                Paragraph(str(tx.get("type", "") or ""), body_style),
                Paragraph(str(tx.get("reference_no", "") or ""), body_style),
                Paragraph(str(tx.get("vehicle", "") or ""), body_style),
                Paragraph(str(tx.get("quantity", 0) or 0), body_style),
                Paragraph(str(tx.get("status", "") or ""), body_style),
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        filename = f"PO_Statement_{po_no}.pdf"
        return StreamingResponse(buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"})
    return {"error": "Unsupported format"}


# ============ BULK IMPORT ROUTES ============

@api_router.get("/import/template/{entity}")
async def get_import_template(entity: str, current_user: dict = Depends(get_download_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    templates = {
        "trucks": {"title": "Trucks Import Template",
            "headers": ["Vehicle Number*", "Transporter Name", "Driver Name", "Driver Mobile",
                        "Helper Name", "Helper Mobile", "Tare Weight (MT)", "Laden Gross Vehicle Weight (MT)"],
            "example": ["MH12AB1234", "ABC Transport", "John Doe", "9876543210",
                        "Helper Name", "9876543211", "5.5", "10"]},
        "products": {"title": "Products Import Template",
            "headers": ["Product Name*", "Product Code*", "Category", "HSN Code", "Unit", "Description"],
            "example": ["Cement", "CEM001", "Building Materials", "2523", "MT", "Portland Cement"]},
        "companies": {"title": "Companies Import Template",
            "headers": ["Company Name*", "Address", "City", "State", "Country", "PIN Code",
                        "Contact Person", "Phone", "Email"],
            "example": ["ABC Corp", "123 Main St", "Mumbai", "Maharashtra", "India", "400001",
                        "John Doe", "9876543210", "contact@abc.com"]},
        "transporters": {"title": "Transporters Import Template",
            "headers": ["Transporter Name*", "Contact Person", "Phone", "Email", "Address", "GST Number"],
            "example": ["XYZ Transport", "Manager Name", "9876543210", "xyz@transport.com",
                        "456 Transport Hub", "27XXXXX1234X1Z5"]}
    }

    if entity not in templates:
        raise HTTPException(status_code=400, detail=f"Unknown entity: {entity}")

    template = templates[entity]
    ws.title = template["title"]
    for col, header in enumerate(template["headers"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    for col, value in enumerate(template["example"], 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
    ws.cell(row=4, column=1, value="Instructions:")
    ws.cell(row=5, column=1, value="* = Required field")
    ws.cell(row=6, column=1, value="Delete the example row before importing")
    ws.cell(row=7, column=1, value="Do not modify headers")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"{entity}_import_template.xlsx"
    return StreamingResponse(buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


# A bulk import is a mass create, so it is gated by the same matrix key as
# creating one of that entity by hand.
IMPORT_PERMISSIONS = {
    "trucks": "Trucks (Create)",
    "products": "Products (Create)",
    "companies": "Companies (Create)",
    "transporters": "Transporters (Create)",
}


@api_router.post("/import/{entity}")
async def bulk_import(entity: str, file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    from openpyxl import load_workbook

    permission_key = IMPORT_PERMISSIONS.get(entity)
    if not permission_key:
        raise HTTPException(status_code=400, detail=f"Unsupported import entity: {entity}")
    await check_permission(current_user, permission_key)

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        def is_instruction_row(row_values):
            values = [str(v).strip().lower() for v in row_values if v not in (None, "")]
            if not values:
                return False
            if len(values) == 1:
                text = values[0]
                skip_phrases = ["instructions", "* = required field", "delete the example row before importing",
                                "do not modify headers", "download all documents"]
                if any(phrase in text for phrase in skip_phrases):
                    return True
            return False

        imported = 0
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row) or is_instruction_row(row):
                continue
            row_data = dict(zip(headers, row))
            try:
                if entity == "trucks":
                    vehicle_number = row_data.get("Vehicle Number*")
                    if not vehicle_number:
                        errors.append(f"Row {row_idx}: Vehicle Number is required")
                        continue
                    async with AsyncSessionLocal() as session:
                        stmt = select(sql_models.Truck).where(sql_models.Truck.vehicle_number == vehicle_number)
                        tfilter = tenant_filter(sql_models.Truck)
                        if tfilter is not None:
                            stmt = stmt.where(tfilter)
                        existing = (await session.execute(stmt)).scalar_one_or_none()
                        if existing:
                            errors.append(f"Row {row_idx}: Vehicle {vehicle_number} already exists")
                            continue
                        truck = sql_models.Truck(
                            id=str(uuid.uuid4()),
                            tenant_id=tenant_id_for_current_user(current_user),
                            vehicle_number=vehicle_number,
                            transporter_name=row_data.get("Transporter Name") or "",
                            driver_name=row_data.get("Driver Name") or "",
                            driver_mobile=row_data.get("Driver Mobile") or "",
                            helper_name=row_data.get("Helper Name") or "",
                            helper_mobile=row_data.get("Helper Mobile") or "",
                            tare_weight_mt=float(row_data.get("Tare Weight (MT)") or 0),
                            capacity_mt=float(row_data.get("Laden Gross Vehicle Weight (MT)") or 0),
                            drivers=[],
                            photos=[],
                            created_at=datetime.now(timezone.utc),
                        )
                        session.add(truck)
                        await session.commit()
                    imported += 1
                elif entity == "products":
                    product_name = row_data.get("Product Name*")
                    product_code = row_data.get("Product Code*")
                    if not product_name or not product_code:
                        errors.append(f"Row {row_idx}: Product Name and Code are required")
                        continue
                    async with AsyncSessionLocal() as session:
                        stmt = select(sql_models.Product).where(sql_models.Product.product_code == product_code)
                        tfilter = tenant_filter(sql_models.Product)
                        if tfilter is not None:
                            stmt = stmt.where(tfilter)
                        existing = (await session.execute(stmt)).scalar_one_or_none()
                        if existing:
                            errors.append(f"Row {row_idx}: Product code {product_code} already exists")
                            continue
                        product = sql_models.Product(
                            id=str(uuid.uuid4()),
                            tenant_id=tenant_id_for_current_user(current_user),
                            product_name=product_name,
                            product_code=product_code,
                            category=row_data.get("Category") or "",
                            hsn_code=row_data.get("HSN Code") or "",
                            unit_of_measurement=row_data.get("Unit") or "MT",
                            product_description=row_data.get("Description") or "",
                            assigned_roles=[],
                            created_at=datetime.now(timezone.utc),
                        )
                        session.add(product)
                        await session.commit()
                    imported += 1
                elif entity == "companies":
                    company_name = row_data.get("Company Name*")
                    if not company_name:
                        errors.append(f"Row {row_idx}: Company Name is required")
                        continue
                    async with AsyncSessionLocal() as session:
                        company = sql_models.Company(
                            id=str(uuid.uuid4()),
                            tenant_id=tenant_id_for_current_user(current_user),
                            name=company_name,
                            address=row_data.get("Address") or "",
                            city=row_data.get("City") or "",
                            state=row_data.get("State") or "",
                            country=row_data.get("Country") or "India",
                            pin_code=row_data.get("PIN Code") or "",
                            contact_person_name=row_data.get("Contact Person") or "",
                            telephone=row_data.get("Phone") or "",
                            primary_email=row_data.get("Email") or "",
                            users=[],
                            added_on=datetime.now(timezone.utc),
                            created_at=datetime.now(timezone.utc),
                        )
                        session.add(company)
                        await session.commit()
                    imported += 1
                elif entity == "transporters":
                    transporter_name = row_data.get("Transporter Name*")
                    if not transporter_name:
                        errors.append(f"Row {row_idx}: Transporter Name is required")
                        continue
                    async with AsyncSessionLocal() as session:
                        transporter = sql_models.Transporter(
                            id=str(uuid.uuid4()),
                            tenant_id=tenant_id_for_current_user(current_user),
                            name=transporter_name,
                            contact_person_name=row_data.get("Contact Person") or "",
                            mobile_number=row_data.get("Phone") or "",
                            email=row_data.get("Email") or "",
                            address=row_data.get("Address") or "",
                            gst_number=row_data.get("GST Number") or "",
                            company_ids=[],
                            users=[],
                            created_at=datetime.now(timezone.utc),
                        )
                        session.add(transporter)
                        await session.commit()
                    imported += 1
                else:
                    raise HTTPException(status_code=400, detail=f"Unknown entity: {entity}")
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        return {"success": True, "imported": imported, "errors": errors[:10], "total_errors": len(errors)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to process file: {str(e)}")


# ============ ANALYTICS ROUTES ============

@api_router.get("/analytics/dashboard")
async def get_dashboard_analytics(current_user: dict = Depends(get_current_user)):
    companies_count = await db.companies.count_documents({})
    users_count = await db.users.count_documents({})
    transporters_count = await db.transporters.count_documents({})
    trucks_count = await db.trucks.count_documents({})
    products_count = await db.products.count_documents({})
    depots_count = await db.depots.count_documents({})
    orders_count = await db.delivery_orders.count_documents({})
    liftings_count = await db.liftings.count_documents({})

    query = {}
    if not current_user.get("is_master_admin"):
        query["from_company_id"] = current_user["company_id"]
    query["status"] = "Open"
    open_orders = await db.delivery_orders.count_documents(query)
    query["status"] = "In Progress"
    in_progress = await db.delivery_orders.count_documents(query)
    query["status"] = "Completed"
    completed = await db.delivery_orders.count_documents(query)

    pending_verification = await db.liftings.count_documents({"unloading_status": "Pending"})
    verified = await db.liftings.count_documents({"unloading_status": "Verified"})

    product_filter = await build_product_filter(current_user, "product_id")

    open_dos = await db.delivery_orders.find(
        {**product_filter, "status": {"$in": ["Open", "In Progress"]}},
        {"_id": 0, "id": 1, "do_order_no": 1, "product_name": 1, "product_id": 1,
         "total_quantity_mt": 1, "lifted_quantity_mt": 1, "remaining_quantity_mt": 1, "status": 1}
    ).to_list(100)

    # ---------- Product allocation & distribution logistics metrics ----------

    # Remaining DO quantity by product
    do_remaining_rows = await db.delivery_orders.find(
        {**product_filter, "status": {"$ne": "Completed"}},
        {"_id": 0, "product_id": 1, "product_name": 1, "remaining_quantity_mt": 1}
    ).to_list(1000)
    do_remaining_by_product = {}
    for r in do_remaining_rows:
        pid = r.get("product_id") or "unknown"
        info = do_remaining_by_product.setdefault(pid, {"product_name": "", "remaining_qty": 0})
        info["remaining_qty"] += r.get("remaining_quantity_mt") or 0

    # Available stock by product from depot inventory
    inventory_rows = await db.depot_inventory.find(
        product_filter,
        {"_id": 0, "product_id": 1, "product_name": 1, "available_quantity": 1}
    ).to_list(1000)
    available_stock_by_product = {}
    for r in inventory_rows:
        pid = r.get("product_id") or "unknown"
        info = available_stock_by_product.setdefault(pid, {"product_name": "", "available_stock": 0})
        info["available_stock"] += r.get("available_quantity") or 0

    # Remaining PO quantity by product
    po_remaining_rows = await db.purchase_orders.find(
        {**product_filter, "status": {"$ne": "Completed"}},
        {"_id": 0, "product_id": 1, "product_name": 1, "remaining_quantity_mt": 1}
    ).to_list(1000)
    po_remaining_by_product = {}
    for r in po_remaining_rows:
        pid = r.get("product_id") or "unknown"
        info = po_remaining_by_product.setdefault(pid, {"product_name": "", "remaining_qty": 0})
        info["remaining_qty"] += r.get("remaining_quantity_mt") or 0

    # Dispatch quantity for today, yesterday, and day before yesterday
    today = datetime.now(timezone.utc)
    today_key = today.strftime("%Y-%m-%d")
    yesterday_key = (today - timedelta(days=1)).strftime("%Y-%m-%d")
    day_before_key = (today - timedelta(days=2)).strftime("%Y-%m-%d")
    dispatch_start = (today - timedelta(days=2)).replace(hour=0, minute=0, second=0, microsecond=0)
    dispatch_end = today.replace(hour=23, minute=59, second=59, microsecond=999999)

    dispatch_by_product = {}
    depot_dispatch_by_product = {}

    def add_dispatch(product_id, name, qty, date_key):
        pid = product_id or "unknown"
        if pid not in dispatch_by_product:
            dispatch_by_product[pid] = {"product_name": name or "Unknown", "today": 0, "yesterday": 0, "day_before_yesterday": 0}
        dispatch_by_product[pid][date_key] += qty

    def add_depot_dispatch(product_id, depot_id, depot_name, qty, date_key):
        pid = product_id or "unknown"
        if pid not in depot_dispatch_by_product:
            depot_dispatch_by_product[pid] = {}
        did = depot_id or "unknown"
        if did not in depot_dispatch_by_product[pid]:
            depot_dispatch_by_product[pid][did] = {
                "depot_id": did,
                "depot_name": depot_name or "Unknown Depot",
                "dispatch_today_qty": 0,
                "dispatch_yesterday_qty": 0,
                "dispatch_day_before_yesterday_qty": 0
            }
        depot_dispatch_by_product[pid][did][date_key] += qty

    secondary_dispatches = await db.liftings.find(
        {
            **product_filter,
            "lifting_type": "Secondary",
            "date_of_loading": {"$gte": dispatch_start, "$lte": dispatch_end}
        },
        {"_id": 0, "product_id": 1, "product_name": 1, "quantity_mt": 1,
         "date_of_loading": 1, "loading_point_id": 1, "loading_point_name": 1}
    ).to_list(1000)

    for item in secondary_dispatches:
        raw = item.get("date_of_loading")
        date_value = (raw or "")[:10] if raw else ""
        qty = item.get("quantity_mt") or 0
        if date_value == today_key:
            add_dispatch(item.get("product_id"), item.get("product_name"), qty, "today")
            add_depot_dispatch(item.get("product_id"), item.get("loading_point_id"), item.get("loading_point_name"), qty, "dispatch_today_qty")
        elif date_value == yesterday_key:
            add_dispatch(item.get("product_id"), item.get("product_name"), qty, "yesterday")
            add_depot_dispatch(item.get("product_id"), item.get("loading_point_id"), item.get("loading_point_name"), qty, "dispatch_yesterday_qty")
        elif date_value == day_before_key:
            add_dispatch(item.get("product_id"), item.get("product_name"), qty, "day_before_yesterday")
            add_depot_dispatch(item.get("product_id"), item.get("loading_point_id"), item.get("loading_point_name"), qty, "dispatch_day_before_yesterday_qty")

    pickup_dispatches = await db.pickups.find(
        {
            **product_filter,
            "status": "verified",
            "verified_at": {"$gte": dispatch_start, "$lte": dispatch_end}
        },
        {"_id": 0, "product_id": 1, "product_name": 1, "weight_mt": 1,
         "verified_at": 1, "source_id": 1, "source_name": 1}
    ).to_list(1000)

    for item in pickup_dispatches:
        raw = item.get("verified_at")
        date_value = (raw or "")[:10] if raw else ""
        qty = item.get("weight_mt") or 0
        if date_value == today_key:
            add_dispatch(item.get("product_id"), item.get("product_name"), qty, "today")
            add_depot_dispatch(item.get("product_id"), item.get("source_id"), item.get("source_name"), qty, "dispatch_today_qty")
        elif date_value == yesterday_key:
            add_dispatch(item.get("product_id"), item.get("product_name"), qty, "yesterday")
            add_depot_dispatch(item.get("product_id"), item.get("source_id"), item.get("source_name"), qty, "dispatch_yesterday_qty")
        elif date_value == day_before_key:
            add_dispatch(item.get("product_id"), item.get("product_name"), qty, "day_before_yesterday")
            add_depot_dispatch(item.get("product_id"), item.get("source_id"), item.get("source_name"), qty, "dispatch_day_before_yesterday_qty")

    # Stock by depot per product (merged with depot-level dispatch)
    inventory_details = await db.depot_inventory.find(
        product_filter,
        {"_id": 0, "product_id": 1, "depot_id": 1, "depot_name": 1, "available_quantity": 1}
    ).to_list(1000)
    stock_by_depot_by_product = {}
    for item in inventory_details:
        pid = item.get("product_id") or "unknown"
        depot_id = item.get("depot_id") or "unknown"
        depot_dispatch = depot_dispatch_by_product.get(pid, {}).get(depot_id, {})
        stock_by_depot_by_product.setdefault(pid, []).append({
            "depot_id": depot_id,
            "depot_name": item.get("depot_name") or "Unknown Depot",
            "available_quantity": item.get("available_quantity") or 0,
            "dispatch_today_qty": depot_dispatch.get("dispatch_today_qty", 0),
            "dispatch_yesterday_qty": depot_dispatch.get("dispatch_yesterday_qty", 0),
            "dispatch_day_before_yesterday_qty": depot_dispatch.get("dispatch_day_before_yesterday_qty", 0)
        })

    # DO total quantity by product (for liftings_product_wise X/Y format)
    do_total_rows = await db.delivery_orders.find(
        product_filter,
        {"_id": 0, "product_id": 1, "product_name": 1, "total_quantity_mt": 1}
    ).to_list(1000)
    do_total_by_product = {}
    for r in do_total_rows:
        pid = r.get("product_id") or "unknown"
        info = do_total_by_product.setdefault(pid, {"product_name": "", "total_qty": 0})
        info["total_qty"] += r.get("total_quantity_mt") or 0

    # Primary / Secondary liftings breakdown (per product)
    primary_liftings = await db.liftings.find(
        {"lifting_type": "Primary"},
        {"_id": 0, "product_id": 1, "product_name": 1, "quantity_mt": 1, "unloading_point_type": 1}
    ).to_list(2000)
    secondary_liftings = await db.liftings.find(
        {"lifting_type": "Secondary"},
        {"_id": 0, "product_id": 1, "product_name": 1, "quantity_mt": 1, "unloading_point_type": 1}
    ).to_list(2000)

    primary_to_depot = [l for l in primary_liftings if l.get("unloading_point_type") == "Depot"]
    primary_to_client = [l for l in primary_liftings if l.get("unloading_point_type") != "Depot"]
    secondary_to_company = [l for l in secondary_liftings if l.get("unloading_point_type") == "Company"]
    secondary_to_depot = [l for l in secondary_liftings if l.get("unloading_point_type") == "Depot"]

    primary_by_product = {}
    for l in primary_liftings:
        pid = l.get("product_id") or "unknown"
        info = primary_by_product.setdefault(pid, {"product_name": l.get("product_name") or "Unknown", "lifted_qty": 0, "to_depot": 0, "to_client": 0})
        qty = l.get("quantity_mt") or 0
        info["lifted_qty"] += qty
        if l.get("unloading_point_type") == "Depot":
            info["to_depot"] += qty
        else:
            info["to_client"] += qty

    secondary_by_product = {}
    for l in secondary_liftings:
        pid = l.get("product_id") or "unknown"
        info = secondary_by_product.setdefault(pid, {"product_name": l.get("product_name") or "Unknown", "lifted_qty": 0, "to_company": 0, "to_depot": 0})
        qty = l.get("quantity_mt") or 0
        info["lifted_qty"] += qty
        if l.get("unloading_point_type") == "Company":
            info["to_company"] += qty
        elif l.get("unloading_point_type") == "Depot":
            info["to_depot"] += qty

    liftings_product_wise = []
    all_product_ids = set(list(do_total_by_product.keys()) + list(primary_by_product.keys()) + list(secondary_by_product.keys()))
    for pid in all_product_ids:
        do_info = do_total_by_product.get(pid, {"product_name": "Unknown", "total_qty": 0})
        primary_info = primary_by_product.get(pid, {"product_name": do_info["product_name"], "lifted_qty": 0, "to_depot": 0, "to_client": 0})
        secondary_info = secondary_by_product.get(pid, {"product_name": do_info["product_name"], "lifted_qty": 0, "to_company": 0, "to_depot": 0})
        liftings_product_wise.append({
            "product_id": pid,
            "product_name": primary_info["product_name"] or secondary_info["product_name"] or do_info["product_name"],
            "do_total_qty": round(do_info["total_qty"], 2),
            "primary_lifted": round(primary_info["lifted_qty"], 2),
            "primary_to_depot": round(primary_info["to_depot"], 2),
            "primary_to_client": round(primary_info["to_client"], 2),
            "secondary_lifted": round(secondary_info["lifted_qty"], 2),
            "secondary_to_company": round(secondary_info["to_company"], 2),
            "secondary_to_depot": round(secondary_info["to_depot"], 2)
        })

    # Pending verification breakdown by product
    pending_liftings = await db.liftings.find(
        {"unloading_status": "Pending"},
        {"_id": 0, "product_id": 1, "product_name": 1, "quantity_mt": 1, "lifting_type": 1}
    ).to_list(1000)
    pending_by_product = {}
    for l in pending_liftings:
        pid = l.get("product_id") or "unknown"
        info = pending_by_product.setdefault(pid, {"product_name": l.get("product_name") or "Unknown", "total_qty": 0, "count": 0})
        info["total_qty"] += l.get("quantity_mt") or 0
        info["count"] += 1

    # Company deliveries breakdown (product-wise with ledger)
    company_deliveries = await db.liftings.find(
        {"unloading_point_type": "Company"},
        {"_id": 0, "product_id": 1, "product_name": 1, "quantity_mt": 1, "net_weight_mt": 1,
         "lifting_type": 1, "unloading_status": 1, "lifting_no": 1, "unloading_point_name": 1,
         "date_of_loading": 1, "date_of_unloading": 1, "vehicle_number": 1}
    ).sort("date_of_loading", -1).to_list(500)

    company_deliveries_by_product = {}
    for l in company_deliveries:
        pid = l.get("product_id") or "unknown"
        qty = l.get("net_weight_mt") or l.get("quantity_mt") or 0
        info = company_deliveries_by_product.setdefault(pid, {
            "product_name": l.get("product_name") or "Unknown",
            "total_qty": 0, "count": 0, "verified_qty": 0, "pending_qty": 0, "liftings": []
        })
        info["total_qty"] += qty
        info["count"] += 1
        if l.get("unloading_status") == "Verified":
            info["verified_qty"] += qty
        else:
            info["pending_qty"] += qty
        info["liftings"].append({
            "lifting_no": l.get("lifting_no"),
            "company_name": l.get("unloading_point_name"),
            "quantity": qty,
            "lifting_type": l.get("lifting_type"),
            "status": l.get("unloading_status"),
            "date_of_loading": l.get("date_of_loading"),
            "date_of_unloading": l.get("date_of_unloading"),
            "vehicle_number": l.get("vehicle_number")
        })

    # Product allocation metrics (all products, not just those with activity)
    all_products = await db.products.find(
        await build_product_filter(current_user, "id"),
        {"_id": 0, "id": 1, "product_name": 1}
    ).to_list(500)
    product_name_lookup = {p.get("id"): p.get("product_name") for p in all_products}

    product_metrics = []
    for pid in product_name_lookup.keys():
        do_info = do_remaining_by_product.get(pid, {"product_name": "", "remaining_qty": 0})
        stock_info = available_stock_by_product.get(pid, {"product_name": "", "available_stock": 0})
        po_info = po_remaining_by_product.get(pid, {"product_name": "", "remaining_qty": 0})
        dispatch_info = dispatch_by_product.get(pid, {"product_name": "", "today": 0, "yesterday": 0, "day_before_yesterday": 0})

        product_name = (
            product_name_lookup.get(pid) or
            do_info.get("product_name") or
            stock_info.get("product_name") or
            po_info.get("product_name") or
            dispatch_info.get("product_name") or
            "Unknown"
        )

        product_metrics.append({
            "product_id": pid,
            "product_name": product_name,
            "remaining_do_qty": round(do_info["remaining_qty"], 2),
            "available_stock_qty": round(stock_info["available_stock"], 2),
            "remaining_po_qty": round(po_info["remaining_qty"], 2),
            "dispatch_today_qty": round(dispatch_info["today"], 2),
            "dispatch_yesterday_qty": round(dispatch_info["yesterday"], 2),
            "dispatch_day_before_yesterday_qty": round(dispatch_info["day_before_yesterday"], 2),
            "stock_by_depot": stock_by_depot_by_product.get(pid, [])
        })

    return {
        "counts": {"companies": companies_count, "users": users_count, "transporters": transporters_count,
                   "trucks": trucks_count, "products": products_count, "depots": depots_count,
                   "delivery_orders": orders_count, "liftings": liftings_count},
        "orders_by_status": {"open": open_orders, "in_progress": in_progress, "completed": completed},
        "liftings_by_status": {"pending": pending_verification, "verified": verified},
        "open_delivery_orders": open_dos,
        "liftings_product_wise": liftings_product_wise,
        "product_metrics": product_metrics,
        "pending_by_product": [{"product_id": k, **v} for k, v in pending_by_product.items()],
        "primary_summary": {
            "total_count": len(primary_liftings),
            "total_qty": round(sum(l.get("quantity_mt", 0) for l in primary_liftings), 2),
            "to_depot_count": len(primary_to_depot),
            "to_depot_qty": round(sum(l.get("quantity_mt", 0) for l in primary_to_depot), 2),
            "to_client_count": len(primary_to_client),
            "to_client_qty": round(sum(l.get("quantity_mt", 0) for l in primary_to_client), 2)
        },
        "secondary_summary": {
            "total_count": len(secondary_liftings),
            "total_qty": round(sum(l.get("quantity_mt", 0) for l in secondary_liftings), 2),
            "to_company_count": len(secondary_to_company),
            "to_company_qty": round(sum(l.get("quantity_mt", 0) for l in secondary_to_company), 2),
            "to_depot_count": len(secondary_to_depot),
            "to_depot_qty": round(sum(l.get("quantity_mt", 0) for l in secondary_to_depot), 2)
        },
        "company_deliveries": [{"product_id": k, **v} for k, v in company_deliveries_by_product.items()],
        "company_deliveries_total": {
            "total_qty": round(sum(l.get("quantity_mt") or l.get("net_weight_mt", 0) for l in company_deliveries), 2),
            "total_count": len(company_deliveries),
            "verified_count": len([l for l in company_deliveries if l.get("unloading_status") == "Verified"]),
            "pending_count": len([l for l in company_deliveries if l.get("unloading_status") != "Verified"])
        }
    }


# ============ ROUTE INCLUSION ============

from routes import (
    reports_router, companies_router, transporters_router, trucks_router,
    products_router, railway_sidings_router, railway_zones_router, depots_router,
    delivery_orders_router, liftings_router, permissions_router, product_access_router,
    depot_access_router, purchase_orders_router, pickups_router,
    verified_trucks_router, company_inventory_router, tenants_router,
    source_access_router, sources_router, product_management_router,
    locations_router, leads_router, firms_router, employees_router,
    invoicing_router, payments_router, notes_router, stock_transfers_router
)

api_router.include_router(reports_router)
api_router.include_router(companies_router)
api_router.include_router(transporters_router)
api_router.include_router(trucks_router)
api_router.include_router(products_router)
api_router.include_router(railway_sidings_router)
api_router.include_router(railway_zones_router)
api_router.include_router(depots_router)
api_router.include_router(delivery_orders_router)
api_router.include_router(liftings_router)
api_router.include_router(permissions_router)
api_router.include_router(product_access_router)
api_router.include_router(depot_access_router)
api_router.include_router(purchase_orders_router)
api_router.include_router(pickups_router)
api_router.include_router(verified_trucks_router)
api_router.include_router(company_inventory_router)
api_router.include_router(tenants_router)
api_router.include_router(source_access_router)
api_router.include_router(sources_router)
api_router.include_router(product_management_router)
api_router.include_router(locations_router)
api_router.include_router(leads_router)
api_router.include_router(firms_router)
api_router.include_router(employees_router)
api_router.include_router(invoicing_router)
api_router.include_router(payments_router)
api_router.include_router(notes_router)
api_router.include_router(stock_transfers_router)

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

MASTER_ADMIN_MOBILE = os.environ.get('MASTER_ADMIN_MOBILE')
MASTER_ADMIN_COUNTRY_CODE = os.environ.get('MASTER_ADMIN_COUNTRY_CODE', '91')
MASTER_ADMIN_PASSWORD = os.environ.get('MASTER_ADMIN_PASSWORD')
MASTER_ADMIN_NAME = os.environ.get('MASTER_ADMIN_NAME', 'Master Admin')
MASTER_ADMIN_EMAIL = os.environ.get('MASTER_ADMIN_EMAIL', 'admin@logitrackpro.com')

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()


async def seed_permissions():
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(sql_models.Permission).where(sql_models.Permission.id == "role_permissions")
        )
        existing = result.scalar_one_or_none()
    if existing:
        return
    perm_doc = sql_models.Permission(
        id="role_permissions",
        permissions=PERMISSION_DEFAULTS,
        updated_at=datetime.now(timezone.utc),
    )
    async with AsyncSessionLocal() as session:
        session.add(perm_doc)
        await session.commit()
    logger.info("Default permissions seeded")


async def seed_platform_tenant():
    """Create the platform tenant if the migration has not seeded it yet."""
    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(sql_models.Tenant).where(sql_models.Tenant.id == PLATFORM_TENANT_ID)
        )
        existing = result.scalar_one_or_none()
        if existing:
            return
        session.add(sql_models.Tenant(
            id=PLATFORM_TENANT_ID,
            name="Platform",
            slug="platform",
            status="active",
            subscription_plan="platform",
            branding={"name": "IBRMCO"},
            feature_flags={},
            created_at=datetime.now(timezone.utc),
        ))
        await session.commit()
    logger.info("Platform tenant seeded")


@app.on_event("startup")
async def seed_master_admin():
    await init_db()
    await seed_platform_tenant()
    if not MASTER_ADMIN_MOBILE or not MASTER_ADMIN_PASSWORD:
        logger.warning("Master Admin credentials not configured")
        return
    full_mobile = f"{MASTER_ADMIN_COUNTRY_CODE}{MASTER_ADMIN_MOBILE}"
    async with AsyncSessionLocal() as session:
        result = await session.execute(select(sql_models.User).where(sql_models.User.is_master_admin == True))
        existing = result.scalar_one_or_none()
    if existing:
        if not existing.tenant_id:
            async with AsyncSessionLocal() as session:
                await session.execute(
                    update(sql_models.User).where(sql_models.User.id == existing.id)
                    .values(tenant_id=PLATFORM_TENANT_ID)
                )
                await session.commit()
        async with AsyncSessionLocal() as session:
            await session.execute(
                update(sql_models.User).where(sql_models.User.is_master_admin == True).values(
                    mobile=full_mobile, country_code=MASTER_ADMIN_COUNTRY_CODE,
                    password=hash_password(MASTER_ADMIN_PASSWORD), email=MASTER_ADMIN_EMAIL,
                    name=MASTER_ADMIN_NAME
                )
            )
            await session.commit()
        logger.info(f"Master Admin updated: +{full_mobile}")
    else:
        master_admin = sql_models.User(
            id=str(uuid.uuid4()), tenant_id=PLATFORM_TENANT_ID, name=MASTER_ADMIN_NAME, mobile=full_mobile,
            country_code=MASTER_ADMIN_COUNTRY_CODE, password=hash_password(MASTER_ADMIN_PASSWORD),
            password_set=True, role="Management", email=MASTER_ADMIN_EMAIL, depot_id=None,
            otp_verified=True, is_master_admin=True,
            created_at=datetime.now(timezone.utc),
        )
        async with AsyncSessionLocal() as session:
            session.add(master_admin)
            await session.commit()
        logger.info(f"Master Admin created: +{full_mobile}")
    await seed_permissions()


@app.on_event("shutdown")
async def shutdown_db():
    await engine.dispose()


@app.get("/")
async def root():
    return {"message": "LogiTrack Pro API v2.1 - InfoEIGHT (MySQL)"}