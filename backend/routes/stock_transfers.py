"""Stock transfer routes (Phase 5).

Inter-depot / inter-company transfers. State machine:
Requested -> Approved -> Dispatched -> Received
       \-> Rejected / Cancelled

Every transition appends an audit row.
"""
from fastapi import APIRouter, HTTPException, Depends
from typing import Optional
from pydantic import BaseModel
from datetime import datetime, timezone
import uuid
import json

from .db_compat import db
from auth_utils import get_current_user, check_permission
from extensions.registry import trigger as ext_trigger

router = APIRouter(tags=["Stock Transfers"])

VALID_TRANSITIONS = {
    "Requested": {"Approved", "Rejected", "Cancelled"},
    "Approved": {"Dispatched", "Rejected", "Cancelled"},
    "Dispatched": {"Received"},
}

TERMINAL = {"Received", "Rejected", "Cancelled"}


class StockTransferCreate(BaseModel):
    product_id: str
    quantity_mt: float
    from_type: str  # Depot | Company
    from_id: str
    to_type: str
    to_id: str
    request_notes: Optional[str] = None


class TransitionPayload(BaseModel):
    notes: Optional[str] = None


class ApprovalMatrixPayload(BaseModel):
    entity: str = "stock_transfer"
    product_id: Optional[str] = None
    amount_threshold: Optional[float] = None
    approver_roles: list = []
    active: bool = True


async def _resolve_approver_roles(product_id: str, quantity_mt: float):
    """Most specific matching matrix wins (product-specific, then highest threshold)."""
    matrices = await db.approval_matrices.find(
        {"entity": "stock_transfer", "active": {"$ne": False}}, {"_id": 0}
    ).to_list(1000)
    if not matrices:
        return None
    matching = []
    for m in matrices:
        pid = m.get("product_id")
        thresh = m.get("amount_threshold")
        if pid and pid != product_id:
            continue
        if thresh is not None and quantity_mt < float(thresh):
            continue
        matching.append(m)
    if not matching:
        return None
    matching.sort(
        key=lambda m: (m.get("product_id") is not None, float(m.get("amount_threshold") or 0)),
        reverse=True,
    )
    roles = matching[0].get("approver_roles")
    if isinstance(roles, str):
        import json as _json
        try:
            roles = _json.loads(roles)
        except Exception:
            roles = []
    return roles or []


async def _resolve_party(party_type: str, party_id: str):
    if party_type not in ("Depot", "Company"):
        raise HTTPException(status_code=400, detail="from_type/to_type must be Depot or Company")
    if party_type == "Depot":
        obj = await db.depots.find_one({"id": party_id})
        name = obj.get("name") if obj else None
    else:
        obj = await db.companies.find_one({"id": party_id})
        name = obj.get("name") if obj else None
    if not obj:
        raise HTTPException(status_code=404, detail=f"{party_type} {party_id} not found")
    return name or ""


async def _audit(transfer_id: str, event: str, actor: dict, payload: dict = None):
    await db.stock_transfer_audit.insert_one({
        "id": str(uuid.uuid4()),
        "transfer_id": transfer_id,
        "event": event,
        "actor_id": actor.get("id"),
        "actor_name": actor.get("name"),
        "payload": json.dumps(payload or {}),
        "created_at": datetime.now(timezone.utc).isoformat(),
    })


async def _inventory_for(party_type: str, party_id: str, product_id: str):
    if party_type == "Depot":
        return await db.depot_inventory.find_one({"depot_id": party_id, "product_id": product_id})
    return await db.company_inventory.find_one({"company_id": party_id, "product_id": product_id})


async def _adjust_locked(party_type: str, party_id: str, product_id: str, delta: float):
    coll = db.depot_inventory if party_type == "Depot" else db.company_inventory
    key = {"depot_id": party_id, "product_id": product_id} if party_type == "Depot" else {"company_id": party_id, "product_id": product_id}
    await coll.update_one(key, {"$inc": {"locked_qty": delta}})


async def _lock_source(transfer: dict):
    inv = await _inventory_for(transfer["from_type"], transfer["from_id"], transfer["product_id"])
    available = float(inv.get("available_quantity", 0)) if inv else 0
    locked = float(inv.get("locked_qty", 0)) if inv else 0
    qty = float(transfer["quantity_mt"])
    if available - locked < qty - 1e-9:
        raise HTTPException(
            status_code=400,
            detail=f"Insufficient stock at source (available {available - locked:.2f} MT, requested {qty:.2f} MT)",
        )
    await _adjust_locked(transfer["from_type"], transfer["from_id"], transfer["product_id"], qty)


async def _unlock_source(transfer: dict):
    await _adjust_locked(transfer["from_type"], transfer["from_id"], transfer["product_id"], -float(transfer["quantity_mt"]))


async def _move_inventory(transfer: dict):
    # Decrement source (available), increment destination. Uses the atomic
    # helpers from liftings.py so the pattern stays consistent.
    from routes.liftings import update_depot_inventory, update_company_inventory

    product = await db.products.find_one({"id": transfer["product_id"]})
    product_code = product.get("product_code") if product else ""
    product_name = transfer.get("product_name") or (product.get("product_name") if product else "")
    qty = float(transfer["quantity_mt"])

    # Source decrement
    if transfer["from_type"] == "Depot":
        # Company id for depot inventory seed: try to keep the existing one.
        inv = await _inventory_for("Depot", transfer["from_id"], transfer["product_id"])
        company_id = inv.get("company_id") if inv else None
        await update_depot_inventory(
            depot_id=transfer["from_id"],
            depot_name=transfer["from_name"],
            product_id=transfer["product_id"],
            product_name=product_name,
            product_code=product_code,
            quantity_change=qty,
            is_incoming=False,
            company_id=company_id or "",
        )
    else:
        await update_company_inventory(
            company_id=transfer["from_id"],
            company_name=transfer["from_name"],
            product_id=transfer["product_id"],
            product_name=product_name,
            product_code=product_code,
            quantity_change=qty,
            is_incoming=False,
        )

    # Destination increment
    if transfer["to_type"] == "Depot":
        inv = await _inventory_for("Depot", transfer["to_id"], transfer["product_id"])
        company_id = inv.get("company_id") if inv else None
        # Fallback: depot's company
        if not company_id:
            depot = await db.depots.find_one({"id": transfer["to_id"]})
            company_id = depot.get("company_id") if depot else ""
        await update_depot_inventory(
            depot_id=transfer["to_id"],
            depot_name=transfer["to_name"],
            product_id=transfer["product_id"],
            product_name=product_name,
            product_code=product_code,
            quantity_change=qty,
            is_incoming=True,
            company_id=company_id or "",
        )
    else:
        await update_company_inventory(
            company_id=transfer["to_id"],
            company_name=transfer["to_name"],
            product_id=transfer["product_id"],
            product_name=product_name,
            product_code=product_code,
            quantity_change=qty,
            is_incoming=True,
        )


def _transfer_or_404(transfer: dict | None):
    if not transfer:
        raise HTTPException(status_code=404, detail="Stock transfer not found")
    return transfer


@router.post("/stock-transfers")
async def create_stock_transfer(data: StockTransferCreate, current_user: dict = Depends(get_current_user)):
    await check_permission(current_user, "Stock Transfers (Create)")

    if data.quantity_mt <= 0:
        raise HTTPException(status_code=400, detail="quantity_mt must be positive")
    if data.from_type == data.to_type and data.from_id == data.to_id:
        raise HTTPException(status_code=400, detail="Source and destination must differ")

    product = await db.products.find_one({"id": data.product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    await ext_trigger("pre_create:stock_transfers", {"data": data.model_dump(), "user": current_user})

    from_name = await _resolve_party(data.from_type, data.from_id)
    to_name = await _resolve_party(data.to_type, data.to_id)

    count = await db.stock_transfers.count_documents({})
    transfer_no = f"TRF-{str(count + 1).zfill(6)}"
    now = datetime.now(timezone.utc).isoformat()
    transfer_id = str(uuid.uuid4())

    transfer = {
        "id": transfer_id,
        "transfer_no": transfer_no,
        "product_id": data.product_id,
        "product_name": product.get("product_name") or "",
        "quantity_mt": float(data.quantity_mt),
        "from_type": data.from_type,
        "from_id": data.from_id,
        "from_name": from_name,
        "to_type": data.to_type,
        "to_id": data.to_id,
        "to_name": to_name,
        "status": "Requested",
        "requested_by": current_user.get("id"),
        "requested_by_name": current_user.get("name"),
        "requested_at": now,
        "request_notes": data.request_notes,
        "created_at": now,
    }
    await db.stock_transfers.insert_one(transfer)
    await _audit(transfer_id, "Requested", current_user, {"quantity_mt": data.quantity_mt})
    await ext_trigger("post_create:stock_transfers", {"transfer": transfer, "user": current_user})

    # Reserve stock at the source so it cannot be consumed elsewhere.
    try:
        await _lock_source(transfer)
    except HTTPException:
        await db.stock_transfers.delete_one({"id": transfer_id})
        await db.stock_transfer_audit.delete_many({"transfer_id": transfer_id})
        raise

    return transfer


@router.get("/stock-transfers")
async def list_stock_transfers(
    status: Optional[str] = None,
    product_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    await check_permission(current_user, "Stock Transfers (View)")
    query = {}
    if status:
        query["status"] = status
    if product_id:
        query["product_id"] = product_id
    return await db.stock_transfers.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)


@router.get("/stock-transfers/export")
async def export_stock_transfers(current_user: dict = Depends(get_current_user)):
    await check_permission(current_user, "Stock Transfers (View)")
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io
    from datetime import datetime, timezone

    transfers = await db.stock_transfers.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Transfers"
    headers = ["Transfer No", "Product", "Qty (MT)", "From", "To", "Status", "Requested By", "Created"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    for row_idx, tr in enumerate(transfers, 2):
        ws.cell(row=row_idx, column=1, value=tr.get("transfer_no", "")).border = thin_border
        ws.cell(row=row_idx, column=2, value=tr.get("product_name", "")).border = thin_border
        ws.cell(row=row_idx, column=3, value=tr.get("quantity_mt", 0)).border = thin_border
        ws.cell(row=row_idx, column=4, value=f"{tr.get('from_type','')} {tr.get('from_name','')}").border = thin_border
        ws.cell(row=row_idx, column=5, value=f"{tr.get('to_type','')} {tr.get('to_name','')}").border = thin_border
        ws.cell(row=row_idx, column=6, value=tr.get("status", "")).border = thin_border
        ws.cell(row=row_idx, column=7, value=tr.get("requested_by_name", "")).border = thin_border
        ws.cell(row=row_idx, column=8, value=str(tr.get("created_at", ""))[:19]).border = thin_border
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"stock_transfers_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/stock-transfers/{transfer_id}")
async def get_stock_transfer(transfer_id: str, current_user: dict = Depends(get_current_user)):
    await check_permission(current_user, "Stock Transfers (View)")
    transfer = await db.stock_transfers.find_one({"id": transfer_id}, {"_id": 0})
    _transfer_or_404(transfer)
    audit = await db.stock_transfer_audit.find({"transfer_id": transfer_id}, {"_id": 0}).sort("created_at", 1).to_list(1000)
    transfer["audit"] = audit
    return transfer


async def _transition(transfer_id: str, target_status: str, event: str, actor: dict, notes_field: str = None, notes: str = None):
    transfer = await db.stock_transfers.find_one({"id": transfer_id})
    _transfer_or_404(transfer)
    current = transfer.get("status")
    if current in TERMINAL:
        raise HTTPException(status_code=400, detail=f"Transfer already {current}")
    allowed = VALID_TRANSITIONS.get(current, set())
    if target_status not in allowed:
        raise HTTPException(status_code=400, detail=f"Cannot transition from {current} to {target_status}")

    now = datetime.now(timezone.utc).isoformat()
    update = {"status": target_status}
    if target_status == "Approved":
        update["approved_by"] = actor.get("id")
        update["approved_by_name"] = actor.get("name")
        update["approved_at"] = now
        if notes:
            update["approval_notes"] = notes
    elif target_status == "Dispatched":
        update["dispatched_by"] = actor.get("id")
        update["dispatched_by_name"] = actor.get("name")
        update["dispatched_at"] = now
        if notes:
            update["dispatch_notes"] = notes
    elif target_status == "Received":
        update["received_by"] = actor.get("id")
        update["received_by_name"] = actor.get("name")
        update["received_at"] = now
        if notes:
            update["receive_notes"] = notes
    elif target_status in ("Rejected", "Cancelled"):
        if notes:
            update["approval_notes"] = notes

    if notes_field and notes and notes_field not in update:
        update[notes_field] = notes

    await db.stock_transfers.update_one({"id": transfer_id}, {"$set": update})
    await _audit(transfer_id, event, actor, {"from": current, "to": target_status, "notes": notes})
    return await db.stock_transfers.find_one({"id": transfer_id}, {"_id": 0})


@router.post("/stock-transfers/{transfer_id}/approve")
async def approve_transfer(transfer_id: str, payload: TransitionPayload = None, current_user: dict = Depends(get_current_user)):
    await check_permission(current_user, "Stock Transfers (Approve)")
    transfer = await db.stock_transfers.find_one({"id": transfer_id})
    _transfer_or_404(transfer)
    if transfer.get("requested_by") == current_user.get("id"):
        raise HTTPException(status_code=400, detail="Requester cannot approve own transfer")
    approver_roles = await _resolve_approver_roles(transfer.get("product_id"), float(transfer.get("quantity_mt") or 0))
    if approver_roles is not None and current_user.get("role") not in approver_roles and not current_user.get("is_master_admin"):
        raise HTTPException(status_code=403, detail="Your role cannot approve this transfer")
    notes = payload.notes if payload else None
    return await _transition(transfer_id, "Approved", "Approved", current_user, notes=notes)


@router.post("/stock-transfers/{transfer_id}/dispatch")
async def dispatch_transfer(transfer_id: str, payload: TransitionPayload = None, current_user: dict = Depends(get_current_user)):
    await check_permission(current_user, "Stock Transfers (Update)")
    notes = payload.notes if payload else None
    return await _transition(transfer_id, "Dispatched", "Dispatched", current_user, notes=notes)


@router.post("/stock-transfers/{transfer_id}/receive")
async def receive_transfer(transfer_id: str, payload: TransitionPayload = None, current_user: dict = Depends(get_current_user)):
    await check_permission(current_user, "Stock Transfers (Update)")
    notes = payload.notes if payload else None
    transfer = await db.stock_transfers.find_one({"id": transfer_id})
    _transfer_or_404(transfer)
    result = await _transition(transfer_id, "Received", "Received", current_user, notes=notes)
    # Release the reservation and move the stock atomically.
    await _unlock_source(transfer)
    await _move_inventory(transfer)
    return result


@router.post("/stock-transfers/{transfer_id}/reject")
async def reject_transfer(transfer_id: str, payload: TransitionPayload = None, current_user: dict = Depends(get_current_user)):
    await check_permission(current_user, "Stock Transfers (Approve)")
    transfer = await db.stock_transfers.find_one({"id": transfer_id})
    _transfer_or_404(transfer)
    if transfer.get("status") not in ("Requested", "Approved"):
        raise HTTPException(status_code=400, detail="Only Requested or Approved transfers can be rejected")
    notes = payload.notes if payload else None
    if not notes:
        raise HTTPException(status_code=400, detail="Rejection reason is required")
    result = await _transition(transfer_id, "Rejected", "Rejected", current_user, notes=notes)
    await _unlock_source(transfer)
    return result


@router.post("/stock-transfers/{transfer_id}/cancel")
async def cancel_transfer(transfer_id: str, payload: TransitionPayload = None, current_user: dict = Depends(get_current_user)):
    await check_permission(current_user, "Stock Transfers (Update)")
    notes = payload.notes if payload else None
    transfer = await db.stock_transfers.find_one({"id": transfer_id})
    _transfer_or_404(transfer)
    if transfer.get("requested_by") != current_user.get("id") and not current_user.get("is_master_admin") and current_user.get("role") != "Management":
        raise HTTPException(status_code=403, detail="Only the requester or Management can cancel")
    result = await _transition(transfer_id, "Cancelled", "Cancelled", current_user, notes=notes)
    await _unlock_source(transfer)
    return result


@router.get("/stock-transfers/{transfer_id}/audit")
async def get_transfer_audit(transfer_id: str, current_user: dict = Depends(get_current_user)):
    await check_permission(current_user, "Stock Transfers (View)")
    await _transfer_or_404(await db.stock_transfers.find_one({"id": transfer_id}))
    return await db.stock_transfer_audit.find({"transfer_id": transfer_id}, {"_id": 0}).sort("created_at", 1).to_list(1000)


# ============ APPROVAL MATRICES ============

@router.get("/approval-matrices")
async def list_approval_matrices(current_user: dict = Depends(get_current_user)):
    await check_permission(current_user, "Stock Transfers (Approve)")
    return await db.approval_matrices.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)


@router.post("/approval-matrices")
async def create_approval_matrix(data: ApprovalMatrixPayload, current_user: dict = Depends(get_current_user)):
    await check_permission(current_user, "Stock Transfers (Approve)")
    if current_user.get("role") != "Management" and not current_user.get("is_master_admin"):
        raise HTTPException(status_code=403, detail="Only Management can manage approval matrices")
    row = {
        "id": str(uuid.uuid4()),
        "entity": data.entity,
        "product_id": data.product_id,
        "amount_threshold": data.amount_threshold,
        "approver_roles": data.approver_roles or [],
        "active": data.active,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.approval_matrices.insert_one(row)
    return row


@router.put("/approval-matrices/{matrix_id}")
async def update_approval_matrix(matrix_id: str, data: ApprovalMatrixPayload, current_user: dict = Depends(get_current_user)):
    await check_permission(current_user, "Stock Transfers (Approve)")
    if current_user.get("role") != "Management" and not current_user.get("is_master_admin"):
        raise HTTPException(status_code=403, detail="Only Management can manage approval matrices")
    existing = await db.approval_matrices.find_one({"id": matrix_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Approval matrix not found")
    await db.approval_matrices.update_one({"id": matrix_id}, {"$set": data.model_dump()})
    return await db.approval_matrices.find_one({"id": matrix_id}, {"_id": 0})


@router.delete("/approval-matrices/{matrix_id}")
async def delete_approval_matrix(matrix_id: str, current_user: dict = Depends(get_current_user)):
    await check_permission(current_user, "Stock Transfers (Approve)")
    if current_user.get("role") != "Management" and not current_user.get("is_master_admin"):
        raise HTTPException(status_code=403, detail="Only Management can manage approval matrices")
    result = await db.approval_matrices.delete_one({"id": matrix_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Approval matrix not found")
    return {"message": "Approval matrix deleted"}
