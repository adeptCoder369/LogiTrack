"""Invoice routes + invoicing engine (Phase 4).

An invoice is generated from a purchase order: header from the PO (client,
billing parent, source), one line from the product. Line rate resolves from
company_pricing (0 when absent - editable inline), GST is invoice-level.
Status flow: Draft -> Issued -> Partially Paid -> Paid (Overdue derived).
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from typing import List, Optional
from pydantic import BaseModel
from datetime import datetime, timezone, timedelta
import uuid
import io

from .db_compat import db
from auth_utils import get_current_user, get_download_user, check_permission

router = APIRouter(tags=["Invoices"])

INVOICE_STATUSES = ("Draft", "Issued", "Partially Paid", "Paid", "Overdue")
EDITABLE_INVOICE_FIELDS = ("gst_rate", "notes")


async def pick_company_rate(company_id: Optional[str], product_id: Optional[str], db=None):
    """Latest valid company_pricing rate for a product (None when absent)."""
    if not company_id or not product_id:
        return None
    from routes.db_compat import db as _db
    proxy = db or _db
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    rows = await proxy.company_pricing.find(
        {"company_id": company_id, "product_id": product_id},
        {"_id": 0},
    ).sort("created_at", -1).to_list(1000)

    if not rows:
        return None

    for row in rows:
        if row.get("valid_from") and row["valid_from"] > today:
            continue
        if row.get("valid_to") and row["valid_to"] < today:
            continue
        return {"rate": row.get("rate", 0), "tier": row.get("tier")}
    return None


def _next_invoice_no(count: int) -> str:
    return f"INV-{str(count + 1).zfill(6)}"


class InvoiceUpdatePayload(BaseModel):
    gst_rate: Optional[float] = None
    notes: Optional[str] = None
    due_days: Optional[int] = 30


class GeneratePayload(BaseModel):
    gst_rate: float = 0
    due_days: int = 30
    notes: Optional[str] = None


async def _invoice_or_404(invoice_id: str) -> dict:
    inv = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return inv


async def _invoice_items(invoice_id: str) -> List[dict]:
    return await db.invoice_items.find({"invoice_id": invoice_id}, {"_id": 0}).to_list(1000)


async def _invoice_payments(invoice_id: str) -> List[dict]:
    return await db.invoice_payments.find({"invoice_id": invoice_id}, {"_id": 0}).to_list(1000)


async def _paid_total(invoice_id: str) -> float:
    rows = await db.invoice_payments.find({"invoice_id": invoice_id}, {"_id": 0, "amount_allocated": 1}).to_list(1000)
    return round(sum(r.get("amount_allocated") or 0 for r in rows), 2)


async def _credit_total(invoice_id: str) -> float:
    rows = await db.credit_notes.find({"invoice_id": invoice_id, "applied": {"$ne": False}}, {"_id": 0, "amount": 1}).to_list(1000)
    return round(sum(r.get("amount") or 0 for r in rows), 2)


def _effective_status(inv: dict, outstanding: float) -> str:
    status = inv.get("status")
    if status == "Paid":
        return "Paid"
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    due = inv.get("due_date") or "9999-99-99"
    if status in ("Issued", "Partially Paid") and due < today:
        return "Overdue"
    if outstanding > 0 and status == "Issued":
        return "Partially Paid" if inv.get("paid_total") else "Issued"
    return status


async def _decorate(inv: dict) -> dict:
    items = await _invoice_items(inv["id"])
    paid = await _paid_total(inv["id"])
    credits = await _credit_total(inv["id"])
    outstanding = round((inv.get("total_amount") or 0) - paid - credits, 2)
    out = dict(inv)
    out["items"] = items
    out["paid_total"] = paid
    out["credit_total"] = credits
    out["outstanding"] = max(outstanding, 0)
    out["effective_status"] = _effective_status(inv, outstanding)
    return out


# ============ GENERATE / CRUD ============

@router.post("/invoices/generate")
async def generate_invoice(payload: GeneratePayload, po_id: str, current_user: dict = Depends(get_current_user)):
    """Generate an invoice (Draft) from a purchase order."""
    await check_permission(current_user, "Invoices (Generate)")

    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")

    if not po.get("product_id"):
        raise HTTPException(status_code=400, detail="PO has no product to invoice")

    existing_draft = await db.invoices.find_one({"po_id": po_id, "status": "Draft"})
    if existing_draft:
        raise HTTPException(status_code=400, detail="A draft invoice already exists for this PO")

    quantity = float(po.get("dispatched_quantity_mt") or 0) or float(po.get("total_quantity_mt") or 0)
    if quantity <= 0:
        raise HTTPException(status_code=400, detail="PO has no quantity to invoice")

    billing_company_id = po.get("billing_company_id") or po.get("to_company_id")
    billing_company_name = po.get("billing_company_name") or po.get("to_company_name") or ""

    rate_row = await pick_company_rate(billing_company_id, po.get("product_id"))
    rate = rate_row["rate"] if rate_row else 0
    tier = rate_row["tier"] if rate_row else None

    subtotal = round(quantity * rate, 2)
    gst_amount = round(subtotal * float(payload.gst_rate) / 100, 2)
    total = round(subtotal + gst_amount, 2)

    now = datetime.now(timezone.utc).isoformat()
    count = await db.invoices.count_documents({})
    invoice_id = str(uuid.uuid4())

    invoice = {
        "id": invoice_id,
        "invoice_no": _next_invoice_no(count),
        "po_id": po_id,
        "po_number": po.get("po_number"),
        "client_company_id": po.get("to_company_id"),
        "client_company_name": po.get("to_company_name"),
        "billing_company_id": billing_company_id,
        "billing_company_name": billing_company_name,
        "source_type": po.get("source_type"),
        "source_id": po.get("source_id"),
        "source_name": po.get("source_name") or po.get("depot_name"),
        "status": "Draft",
        "invoice_date": None,
        "due_date": None,
        "subtotal": subtotal,
        "gst_rate": float(payload.gst_rate),
        "gst_amount": gst_amount,
        "total_amount": total,
        "currency": "INR",
        "notes": payload.notes,
        "created_by": current_user.get("id"),
        "created_at": now,
    }
    await db.invoices.insert_one(invoice)

    await db.invoice_items.insert_one({
        "id": str(uuid.uuid4()),
        "invoice_id": invoice_id,
        "product_id": po.get("product_id"),
        "product_name": po.get("product_name") or "",
        "description": f"Invoice for PO {po.get('po_number')}",
        "quantity_mt": quantity,
        "rate": rate,
        "amount": subtotal,
        "tier": tier,
        "created_at": now,
    })

    return await _decorate(invoice)


@router.get("/invoices")
async def get_invoices(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    await check_permission(current_user, "Invoices (View)")
    query = {}
    if status:
        query["status"] = status
    invoices = await db.invoices.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [await _decorate(inv) for inv in invoices]


@router.get("/invoices/{invoice_id}")
async def get_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    await check_permission(current_user, "Invoices (View)")
    inv = await _invoice_or_404(invoice_id)
    out = await _decorate(inv)
    out["payments"] = await _invoice_payments(invoice_id)
    out["credit_notes"] = await db.credit_notes.find({"invoice_id": invoice_id}, {"_id": 0}).to_list(1000)
    out["debit_notes"] = await db.debit_notes.find({"invoice_id": invoice_id}, {"_id": 0}).to_list(1000)
    return out


@router.put("/invoices/{invoice_id}")
async def update_invoice(invoice_id: str, payload: InvoiceUpdatePayload, current_user: dict = Depends(get_current_user)):
    """Edit a Draft invoice (or its notes once issued)."""
    await check_permission(current_user, "Invoices (Update)")
    inv = await _invoice_or_404(invoice_id)

    update = {}
    if inv.get("status") == "Draft":
        if payload.gst_rate is not None:
            gst_rate = float(payload.gst_rate)
            subtotal = inv.get("subtotal") or 0
            update["gst_rate"] = gst_rate
            update["gst_amount"] = round(subtotal * gst_rate / 100, 2)
            update["total_amount"] = round(subtotal + update["gst_amount"], 2)
        if payload.notes is not None:
            update["notes"] = payload.notes
    else:
        # Issued invoices: notes only.
        if payload.notes is not None:
            update["notes"] = payload.notes
        if payload.gst_rate is not None:
            raise HTTPException(status_code=400, detail="Invoice already issued; GST is locked")

    if not update:
        raise HTTPException(status_code=400, detail="No editable fields provided")

    await db.invoices.update_one({"id": invoice_id}, {"$set": update})
    return await _decorate(await _invoice_or_404(invoice_id))


@router.post("/invoices/{invoice_id}/issue")
async def issue_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """Draft -> Issued (sets invoice date + due date)."""
    await check_permission(current_user, "Invoices (Issue)")
    inv = await _invoice_or_404(invoice_id)
    if inv.get("status") != "Draft":
        raise HTTPException(status_code=400, detail="Only draft invoices can be issued")

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    due = (datetime.now(timezone.utc) + timedelta(days=30)).strftime("%Y-%m-%d")
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {"status": "Issued", "invoice_date": today, "due_date": due}},
    )
    return await _decorate(await _invoice_or_404(invoice_id))


@router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    await check_permission(current_user, "Invoices (Delete)")
    inv = await _invoice_or_404(invoice_id)
    if inv.get("status") != "Draft":
        raise HTTPException(status_code=400, detail="Only draft invoices can be deleted")
    payments = await _invoice_payments(invoice_id)
    if payments:
        raise HTTPException(status_code=400, detail="Invoice has payments; cannot delete")
    await db.invoice_items.delete_many({"invoice_id": invoice_id})
    await db.invoices.delete_one({"id": invoice_id})
    return {"message": "Invoice deleted"}


# ============ EXPORT (PDF / EXCEL) ============

@router.get("/invoices/{invoice_id}/export")
async def export_invoice(invoice_id: str, format: str = "pdf", current_user: dict = Depends(get_download_user)):
    """Invoice as a reportlab PDF or openpyxl Excel workbook."""
    inv = await _invoice_or_404(invoice_id)
    items = await _invoice_items(invoice_id)
    paid = await _paid_total(invoice_id)
    credits = await _credit_total(invoice_id)
    outstanding = round((inv.get("total_amount") or 0) - paid - credits, 2)

    def _rows():
        return [
            [i.get("product_name") or "", i.get("description") or "", i.get("quantity_mt") or 0, i.get("rate") or 0, i.get("amount") or 0]
            for i in items
        ]

    if format == "pdf":
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        doc = SimpleDocTemplate(io.BytesIO(), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph(f"Invoice {inv.get('invoice_no')}", styles['Title']))
        story.append(Spacer(1, 6))
        info = [
            f"Client: {inv.get('client_company_name') or ''}",
            f"Billing: {inv.get('billing_company_name') or ''}",
            f"Source: {inv.get('source_name') or ''}",
            f"PO: {inv.get('po_number') or ''}",
            f"Date: {inv.get('invoice_date') or ''}   Due: {inv.get('due_date') or ''}",
            f"Status: {inv.get('status') or ''}   Outstanding: {outstanding}",
        ]
        for line in info:
            story.append(Paragraph(line, styles['Normal']))
        story.append(Spacer(1, 10))

        data = [["Product", "Description", "Qty (MT)", "Rate", "Amount"]] + _rows()
        table = Table(data, colWidths=[60 * mm, 60 * mm, 25 * mm, 25 * mm, 25 * mm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        story.append(table)
        story.append(Spacer(1, 10))
        totals = [
            f"Subtotal: {inv.get('subtotal') or 0}",
            f"GST ({inv.get('gst_rate') or 0}%): {inv.get('gst_amount') or 0}",
            f"Total: {inv.get('total_amount') or 0}",
            f"Paid: {paid}   Credits: {credits}   Outstanding: {outstanding}",
        ]
        for line in totals:
            story.append(Paragraph(line, styles['Normal']))

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        doc.build(story)
        buffer.seek(0)
        filename = f"Invoice_{inv.get('invoice_no')}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    if format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

        ws.cell(row=1, column=1, value=f"Invoice {inv.get('invoice_no')}").font = Font(bold=True, size=14)
        meta = [
            ("Client", inv.get("client_company_name") or ""),
            ("Billing", inv.get("billing_company_name") or ""),
            ("Source", inv.get("source_name") or ""),
            ("PO", inv.get("po_number") or ""),
            ("Date / Due", f"{inv.get('invoice_date') or ''} / {inv.get('due_date') or ''}"),
            ("Status / Outstanding", f"{inv.get('status') or ''} / {outstanding}"),
        ]
        for idx, (label, value) in enumerate(meta, start=2):
            ws.cell(row=idx, column=1, value=f"{label}: {value}")

        headers = ["Product", "Description", "Qty (MT)", "Rate", "Amount"]
        header_row = len(meta) + 3
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for row_idx, item in enumerate(_rows(), start=header_row + 1):
            for col_idx, value in enumerate(item, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        total_row = header_row + len(items) + 1
        totals = [
            f"Subtotal: {inv.get('subtotal') or 0}",
            f"GST ({inv.get('gst_rate') or 0}%): {inv.get('gst_amount') or 0}",
            f"Total: {inv.get('total_amount') or 0}",
            f"Paid: {paid}   Credits: {credits}   Outstanding: {outstanding}",
        ]
        for idx, line in enumerate(totals, start=1):
            ws.cell(row=total_row + idx - 1, column=1, value=line).font = Font(bold=True)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"Invoice_{inv.get('invoice_no')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    raise HTTPException(status_code=400, detail="Unsupported format")
